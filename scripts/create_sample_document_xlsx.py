"""일회성: document_templates/files/ 샘플 xlsx 생성 (저장소에 바이너리 커밋용)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


def main() -> None:
    root = Path(__file__).resolve().parent.parent / "document_templates" / "files"
    root.mkdir(parents=True, exist_ok=True)

    def header(ws, cell: str, text: str, bold: bool = True) -> None:
        ws[cell] = text
        if bold:
            ws[cell].font = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    header(ws, "A2", "제목라벨")
    header(ws, "A3", "작업일")
    header(ws, "A4", "담당")
    header(ws, "A5", "요약")
    wb.save(root / "sample_notice.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "발췌"
    ws["A1"] = "작업계획 발췌(예시·필수 게이트)"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "작업 코드"
    ws["B2"] = ""
    ws["A3"] = "발췌 구간 표기"
    ws["B3"] = ""
    ws["A4"] = "위험요인 요약"
    ws["B4"] = ""
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48
    wb.save(root / "sample_with_gate.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "메모"
    ws["A1"] = "간단 메모(게이트 없음 예시)"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = "제목"
    ws["B2"] = ""
    ws["A3"] = "내용"
    ws["B3"] = ""
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    wb.save(root / "sample_no_gate.xlsx")

    headers = ["순번", "품목코드", "품목명", "규격", "단위", "수량", "비고"]
    wb = Workbook()
    ws = wb.active
    ws.title = "자재입고"
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
    wb.save(root / "sample_material_receipt_table.xlsx")

    sched_headers = ["순번", "작업내용", "착수(계획)", "완료(계획)", "기간", "비고"]
    wb = Workbook()
    ws = wb.active
    ws.title = "공사일정"
    for i, h in enumerate(sched_headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
    wb.save(root / "construction_schedule.xlsx")

    print("Wrote:", sorted(p.name for p in root.glob("*.xlsx")))


if __name__ == "__main__":
    main()
