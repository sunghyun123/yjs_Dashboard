import io
import json
from datetime import date
from urllib.parse import parse_qs, urlparse

import app.services.kakao_oauth as kakao_oauth
from app.core.config import settings


def _iso_today() -> str:
    """기본 /today 조회 창(과거 며칠~미래 며칠) 안에 들어가도록 오늘 날짜 사용."""
    return date.today().isoformat()


def login_as_admin(client, monkeypatch, tmp_path):
    """카카오 OAuth 호출을 가짜로 두고 화이트리스트에 있는 관리자로 세션을 만든다."""
    wl = tmp_path / "kakao_whitelist.json"
    wl.write_text(
        json.dumps(
            {
                "users": [
                    {
                        "kakao_id": "999888777",
                        "user_id": "admin",
                        "user_name": "테스트관리자",
                        "role": "admin",
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setattr(settings, "KAKAO_WHITELIST_PATH", str(wl))
    monkeypatch.setattr(kakao_oauth, "fetch_oauth_token", lambda code: {"access_token": "dummy"})
    monkeypatch.setattr(kakao_oauth, "fetch_kakao_user_id", lambda token: "999888777")

    r1 = client.get("/api/auth/kakao/login?next=/dashboard.html", follow_redirects=False)
    assert r1.status_code == 302
    loc = r1.headers.get("location") or ""
    assert "kauth.kakao.com" in loc
    qs = parse_qs(urlparse(loc).query)
    assert qs.get("state")
    state = qs["state"][0]

    r2 = client.get(f"/api/auth/kakao/callback?code=fake-code&state={state}", follow_redirects=False)
    assert r2.status_code == 302


def test_static_pages_are_served(client):
    page_markers = {
        "/": ["현장 보고 실시간 현황", "오늘만"],
        "/index.html": ["사진 카테고리 선택", "요청 접수"],
        "/dashboard.html": ["일정 수정", "오늘만"],
        "/admin.html": ["요청 반려", "백업데이터 생성 실행"],
    }

    for path in ["/", "/index.html", "/dashboard.html", "/admin.html"]:
        res = client.get(path)
        assert res.status_code == 200
        assert "text/html" in res.headers.get("content-type", "")
        body = res.text
        for marker in page_markers[path]:
            assert marker in body

    board_redirect = client.get("/board.html", follow_redirects=False)
    assert board_redirect.status_code == 307
    assert board_redirect.headers.get("location") == "/dashboard.html"


def test_auth_login_me_logout_flow(client, monkeypatch, tmp_path):
    login_as_admin(client, monkeypatch, tmp_path)

    me_res = client.get("/api/auth/me")
    assert me_res.status_code == 200
    me = me_res.json()
    assert me["user_id"] == "admin"
    assert me["role"] == "admin"

    logout_res = client.post("/api/auth/logout")
    assert logout_res.status_code == 200
    assert logout_res.json()["message"] == "로그아웃되었습니다."

    unauthorized_me = client.get("/api/auth/me")
    assert unauthorized_me.status_code == 401


def test_schedule_create_and_today_read(client, monkeypatch, tmp_path):
    login_as_admin(client, monkeypatch, tmp_path)
    d = _iso_today()

    create_res = client.post(
        "/api/schedules/execute",
        json={
            "action": "create",
            "schedule_data": {
                "date": d,
                "location": "안양",
                "task": "지중화 공사",
                "person": "김대리",
                "details": "기본 동작 테스트",
                "tags": ["주간"],
                "category": "공사일정",
            },
        },
    )
    assert create_res.status_code == 200
    create_body = create_res.json()
    assert create_body["status"] == "success"

    today_res = client.get("/api/schedules/today")
    assert today_res.status_code == 200
    today_body = today_res.json()
    assert today_body["count"] >= 1
    created = next((item for item in today_body["data"] if item["task"] == "지중화 공사"), None)
    assert created is not None
    assert created.get("shift_type") == "주간"


def test_schedule_create_with_missing_optional_fields(client, monkeypatch, tmp_path):
    login_as_admin(client, monkeypatch, tmp_path)
    d = _iso_today()

    create_res = client.post(
        "/api/schedules/execute",
        json={
            "action": "create",
            "schedule_data": {
                "date": d,
                "location": "군포",
                "task": "야간작업",
            },
        },
    )
    assert create_res.status_code == 200
    assert create_res.json()["status"] == "success"

    today_res = client.get("/api/schedules/today")
    rows = today_res.json()["data"]
    inserted = next((row for row in rows if row["task"] == "야간작업"), None)
    assert inserted is not None
    assert inserted["details"] in ["", None]
    assert inserted["category"] == "공사 일정"


def test_memo_and_worker_status_flow(client, monkeypatch, tmp_path):
    login_as_admin(client, monkeypatch, tmp_path)

    memo_res = client.post(
        "/api/schedules/memos",
        json={"content": "현장 메모 테스트", "memo_type": "일반", "visibility": "all"},
    )
    assert memo_res.status_code == 200
    assert memo_res.json()["status"] == "success"

    memo_list_res = client.get("/api/schedules/memos")
    assert memo_list_res.status_code == 200
    memo_items = memo_list_res.json()["data"]
    assert any(item["content"] == "현장 메모 테스트" for item in memo_items)

    status_set_res = client.post(
        "/api/schedules/worker-status",
        json={
            "user_name": "홍길동",
            "status": "외출",
            "location": "현장 A",
            "until_time": "",
            "note": "장비 점검",
        },
    )
    assert status_set_res.status_code == 200
    assert status_set_res.json()["status"] == "success"

    status_get_res = client.get("/api/schedules/worker-status")
    assert status_get_res.status_code == 200
    statuses = status_get_res.json()["data"]
    assert any(item["user_name"] == "홍길동" and item["status"] == "외출" for item in statuses)


def test_admin_request_queue_basic_flow(client, monkeypatch, tmp_path):
    login_as_admin(client, monkeypatch, tmp_path)

    chat_res = client.post(
        "/api/schedules/chat",
        json={"text": "삭제 요청 테스트", "input_category": "delete_request"},
    )
    assert chat_res.status_code == 200
    chat_body = chat_res.json()
    assert chat_body["intent"] == "incomplete"
    assert "관리자" in chat_body["reply_message"]

    req_res = client.get("/api/admin/requests")
    assert req_res.status_code == 200
    requests = req_res.json()["data"]
    assert len(requests) >= 1
    assert any(row["request_type"] == "delete_request" for row in requests)


def test_execute_update_delete_go_to_admin_queue(client, monkeypatch, tmp_path):
    login_as_admin(client, monkeypatch, tmp_path)
    d = _iso_today()

    # Create a base schedule first.
    create_res = client.post(
        "/api/schedules/execute",
        json={
            "action": "create",
            "schedule_data": {
                "date": d,
                "location": "의왕",
                "task": "맨홀 정비",
                "person": "박대리",
                "details": "큐 테스트용 생성",
                "tags": ["주간"],
                "category": "공사일정",
            },
        },
    )
    assert create_res.status_code == 200
    assert create_res.json()["status"] == "success"

    today_res = client.get("/api/schedules/today")
    assert today_res.status_code == 200
    created = next((r for r in today_res.json()["data"] if r["task"] == "맨홀 정비"), None)
    assert created is not None
    schedule_id = created["id"]

    update_queue_res = client.post(
        "/api/schedules/execute",
        json={
            "action": "update",
            "schedule_id": schedule_id,
            "schedule_data": {
                "date": d,
                "location": "의왕",
                "task": "맨홀 정비(수정요청)",
                "person": "박대리",
                "details": "관리자 승인 필요",
                "tags": ["주간"],
                "category": "공사일정",
            },
        },
    )
    assert update_queue_res.status_code == 200
    assert update_queue_res.json()["status"] == "success"
    assert "관리자 승인 요청" in update_queue_res.json()["message"]

    delete_queue_res = client.post(
        "/api/schedules/execute",
        json={"action": "delete", "schedule_id": schedule_id},
    )
    assert delete_queue_res.status_code == 200
    assert delete_queue_res.json()["status"] == "success"
    assert "관리자 승인 요청" in delete_queue_res.json()["message"]

    admin_req_res = client.get("/api/admin/requests")
    assert admin_req_res.status_code == 200
    rows = admin_req_res.json()["data"]
    assert any(row["request_type"] == "update_request" for row in rows)
    assert any(row["request_type"] == "delete_request" for row in rows)


def test_admin_review_approve_update_and_delete(client, monkeypatch, tmp_path):
    login_as_admin(client, monkeypatch, tmp_path)
    d = _iso_today()

    # 1) Create original schedule.
    create_res = client.post(
        "/api/schedules/execute",
        json={
            "action": "create",
            "schedule_data": {
                "date": d,
                "location": "안산",
                "task": "케이블 포설",
                "person": "이대리",
                "details": "원본",
                "tags": ["야간"],
                "category": "공사일정",
            },
        },
    )
    assert create_res.status_code == 200

    base_rows = client.get("/api/schedules/today").json()["data"]
    base = next((r for r in base_rows if r["task"] == "케이블 포설"), None)
    assert base is not None
    schedule_id = base["id"]

    # 2) Queue update request then approve it.
    queue_update_res = client.post(
        "/api/schedules/execute",
        json={
            "action": "update",
            "schedule_id": schedule_id,
            "schedule_data": {
                "date": d,
                "location": "안산",
                "task": "케이블 포설(승인수정)",
                "person": "이대리",
                "details": "관리자 승인 반영",
                "tags": ["야간", "긴급"],
                "category": "공사일정",
            },
        },
    )
    assert queue_update_res.status_code == 200

    req_rows = client.get("/api/admin/requests").json()["data"]
    update_req = next((r for r in req_rows if r["request_type"] == "update_request"), None)
    assert update_req is not None

    approve_update_res = client.post(
        "/api/admin/requests/review",
        json={"request_id": update_req["id"], "decision": "approve", "schedule_id": schedule_id},
    )
    assert approve_update_res.status_code == 200

    updated_rows = client.get("/api/schedules/today").json()["data"]
    updated = next((r for r in updated_rows if r["id"] == schedule_id), None)
    assert updated is not None
    assert updated["task"] == "케이블 포설(승인수정)"

    # 3) Queue delete request then approve it.
    queue_delete_res = client.post(
        "/api/schedules/execute",
        json={"action": "delete", "schedule_id": schedule_id},
    )
    assert queue_delete_res.status_code == 200

    req_rows_after = client.get("/api/admin/requests").json()["data"]
    delete_req = next((r for r in req_rows_after if r["request_type"] == "delete_request"), None)
    assert delete_req is not None

    approve_delete_res = client.post(
        "/api/admin/requests/review",
        json={
            "request_id": delete_req["id"],
            "decision": "approve",
            "schedule_id": schedule_id,
            "reason": "삭제 스모크 테스트",
        },
    )
    assert approve_delete_res.status_code == 200

    final_rows = client.get("/api/schedules/today").json()["data"]
    assert all(r["id"] != schedule_id for r in final_rows)


def test_direct_update_delete_require_auth(client):
    update_res = client.post(
        "/api/schedules/direct-update",
        json={"schedule_id": 1, "schedule_data": {"date": "2026-03-30", "location": "의정부", "task": "점검"}},
    )
    assert update_res.status_code == 401

    delete_res = client.post("/api/schedules/direct-delete", json={"schedule_id": 1})
    assert delete_res.status_code == 401


def test_documents_templates_and_fill_sample(client, tmp_path, monkeypatch):
    """저장소가 OneDrive 등에 있으면 원본 xlsx 읽기가 PermissionError 날 수 있어, 임시 폴더로 검증한다."""
    from openpyxl import Workbook

    from app.core.config import settings

    doc_root = tmp_path / "document_templates"
    files_dir = doc_root / "files"
    xlsx_dir = doc_root / "xlsx"
    files_dir.mkdir(parents=True)
    xlsx_dir.mkdir(parents=True)
    wb = Workbook()
    wb.save(files_dir / "sample_notice.xlsx")
    manifest = {
        "templates": [
            {
                "id": "sample_notice_xlsx",
                "title": "Test Excel",
                "file": "files/sample_notice.xlsx",
                "kind": "xlsx",
                "sheet": None,
                "ai": {"suggest": "테스트용 짧은 지시", "extract": "테스트용 추출 지시"},
                "fields": [
                    {"id": "doc_title", "label": "메일 제목(추천)", "cell": "A1"},
                    {"id": "work_date", "label": "작업일(추천x)", "cell": "B2"},
                ],
            }
        ]
    }
    (xlsx_dir / "templates.json").write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setattr(settings, "DOCUMENT_TEMPLATES_DIR", str(doc_root.resolve()))

    login_as_admin(client, monkeypatch, tmp_path)
    res = client.get("/api/documents/templates")
    assert res.status_code == 200
    body = res.json()
    assert "templates" in body
    assert any(t.get("id") == "sample_notice_xlsx" for t in body["templates"])
    sample = next(t for t in body["templates"] if t.get("id") == "sample_notice_xlsx")
    assert any(f.get("id") == "doc_title" for f in sample.get("fields", []))

    fill_res = client.post(
        "/api/documents/fill",
        json={
            "template_id": "sample_notice_xlsx",
            "values": {"doc_title": "테스트제목", "work_date": "2026-04-06"},
        },
    )
    assert fill_res.status_code == 200
    assert fill_res.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_documents_required_gate_fill_and_list(client, tmp_path, monkeypatch):
    from openpyxl import Workbook

    from app.core.config import settings

    doc_root = tmp_path / "document_templates"
    files_dir = doc_root / "files"
    xlsx_dir = doc_root / "xlsx"
    files_dir.mkdir(parents=True)
    xlsx_dir.mkdir(parents=True)
    wb = Workbook()
    wb.save(files_dir / "gate.xlsx")
    manifest = {
        "templates": [
            {
                "id": "gate_xlsx",
                "title": "Gate Test",
                "file": "files/gate.xlsx",
                "kind": "xlsx",
                "required_gate": ["work_code"],
                "fields": [
                    {"id": "work_code", "label": "코드(추천x)", "cell": "A1"},
                    {"id": "note", "label": "비고(추천)", "cell": "B1"},
                ],
            }
        ]
    }
    (xlsx_dir / "templates.json").write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setattr(settings, "DOCUMENT_TEMPLATES_DIR", str(doc_root.resolve()))

    login_as_admin(client, monkeypatch, tmp_path)
    res = client.get("/api/documents/templates")
    assert res.status_code == 200
    tpl = next(t for t in res.json()["templates"] if t.get("id") == "gate_xlsx")
    assert tpl.get("required_gate") == ["work_code"]

    bad = client.post(
        "/api/documents/fill",
        json={"template_id": "gate_xlsx", "values": {"note": "x"}},
    )
    assert bad.status_code == 400
    assert "필수" in (bad.json().get("detail") or "")

    ok = client.post(
        "/api/documents/fill",
        json={"template_id": "gate_xlsx", "values": {"work_code": "W1", "note": "y"}},
    )
    assert ok.status_code == 200


def test_documents_export_table_xlsx(client, tmp_path, monkeypatch):
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment

    from app.core.config import settings

    doc_root = tmp_path / "document_templates"
    files_dir = doc_root / "files"
    xlsx_dir = doc_root / "xlsx"
    files_dir.mkdir(parents=True)
    xlsx_dir.mkdir(parents=True)
    src_wb = Workbook()
    src_ws = src_wb.active
    src_ws.title = "Data"
    src_ws["A1"] = "열A"
    src_ws["B1"] = "열B"
    src_ws["A2"] = ""
    src_ws["B2"] = ""
    src_ws["B2"].alignment = Alignment(wrap_text=True, vertical="top")
    src_wb.save(files_dir / "placeholder.xlsx")
    manifest = {
        "templates": [
            {
                "id": "tbl_test_xlsx",
                "title": "표 테스트",
                "file": "files/placeholder.xlsx",
                "kind": "xlsx",
                "extract_mode": "table",
                "sheet": "Data",
                "table_start_row": 2,
                "table_columns": [
                    {"id": "a", "header": "열A"},
                    {"id": "b", "header": "열B"},
                ],
                "fields": [],
            }
        ]
    }
    (xlsx_dir / "templates.json").write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setattr(settings, "DOCUMENT_TEMPLATES_DIR", str(doc_root.resolve()))

    login_as_admin(client, monkeypatch, tmp_path)
    res = client.post(
        "/api/documents/export-table",
        json={
            "template_id": "tbl_test_xlsx",
            "rows": [{"a": "1", "b": "hello"}, {"a": "", "b": ""}, {"a": "2", "b": "z"}],
            "values": {"note": "ignored"},
        },
    )
    assert res.status_code == 200
    assert res.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(res.content))
    ws = wb.active
    assert ws["A1"].value == "열A"
    assert ws["B1"].value == "열B"
    assert ws["A2"].value == "1"
    assert ws["B2"].value == "hello"
    assert ws["A3"].value == "2"
    assert ws["B3"].value == "z"
    assert ws["B2"].alignment.wrap_text is True

    bad = client.post("/api/documents/export-table", json={"template_id": "missing_tbl", "rows": []})
    assert bad.status_code == 404


def test_documents_template_mode_flags(client, tmp_path, monkeypatch):
    from openpyxl import Workbook

    from app.core.config import settings

    doc_root = tmp_path / "document_templates"
    files_dir = doc_root / "files"
    xlsx_dir = doc_root / "xlsx"
    files_dir.mkdir(parents=True)
    xlsx_dir.mkdir(parents=True)
    wb = Workbook()
    wb.save(files_dir / "mode.xlsx")
    manifest = {
        "templates": [
            {
                "id": "extract_only_xlsx",
                "title": "Extract Only",
                "file": "files/mode.xlsx",
                "kind": "xlsx",
                "generate_enabled": False,
                "extract_enabled": True,
                "fields": [{"id": "a", "label": "A(추천)", "cell": "A1"}],
            }
        ]
    }
    (xlsx_dir / "templates.json").write_text(json.dumps(manifest, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setattr(settings, "DOCUMENT_TEMPLATES_DIR", str(doc_root.resolve()))

    login_as_admin(client, monkeypatch, tmp_path)
    res = client.get("/api/documents/templates")
    assert res.status_code == 200
    tpl = next(t for t in res.json()["templates"] if t.get("id") == "extract_only_xlsx")
    assert tpl.get("generate_enabled") is False
    assert tpl.get("extract_enabled") is True

    bad_fill = client.post("/api/documents/fill", json={"template_id": "extract_only_xlsx", "values": {"a": "x"}})
    assert bad_fill.status_code == 400
    assert "비활성화" in (bad_fill.json().get("detail") or "")


def test_pwa_manifest_served(client):
    res = client.get("/site.webmanifest")
    assert res.status_code == 200
    data = res.json()
    assert data.get("display") == "standalone"
    assert "start_url" in data
