"""문서 템플릿 로드 및 xlsx/hwpx 채움 (매니페스트 기준)."""
from __future__ import annotations

import io
import json
import logging
import zipfile
from copy import copy
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def required_gate_field_ids(template: Dict[str, Any]) -> List[str]:
    """templates.json 의 required_gate: 필드 id 목록. 비어 있으면 게이트 없음."""
    g = template.get("required_gate")
    if not isinstance(g, list):
        return []
    out: List[str] = []
    for x in g:
        s = str(x).strip()
        if s:
            out.append(s)
    return out


def validate_template_required_gates(template: Dict[str, Any], values: Dict[str, str]) -> None:
    """필수 게이트 필드가 비어 있으면 ValueError."""
    for k in required_gate_field_ids(template):
        v = values.get(k)
        if v is None or str(v).strip() == "":
            raise ValueError(f"필수 항목이 비어 있습니다: {k}")


def field_ai_recommendable(label: str) -> bool:
    """라벨에 (추천x) 계열이 있으면 AI 추천 제외. (추천)만 추천 대상."""
    s = (label or "").strip()
    lower = s.lower()
    if "(추천x)" in s or "(추천X)" in s or "(추천 x)" in lower:
        return False
    return "(추천)" in s


def escape_xml_text(value: str) -> str:
    return (
        (value or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _read_templates_array_from_file(path: Path, label: str) -> List[Dict[str, Any]]:
    if not path.is_file():
        return []
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        logger.error("%s JSON 파싱 실패: %s", path, e)
        raise ValueError(
            f"document_templates/{label} 문법 오류(쉼표·따옴표 확인): 줄 {e.lineno} 근처 — {e.msg}"
        ) from e
    items = raw.get("templates")
    if items is None:
        return []
    if not isinstance(items, list):
        return []
    return [t for t in items if isinstance(t, dict)]


def template_ai_suggest_instruction(template: Dict[str, Any]) -> Optional[str]:
    """템플릿별 문서 생성(추천)용 짧은 지시. 없으면 None."""
    ai = template.get("ai")
    if not isinstance(ai, dict):
        return None
    s = ai.get("suggest")
    if s is None:
        return None
    t = str(s).strip()
    return t or None


def template_ai_extract_instruction(template: Dict[str, Any]) -> Optional[str]:
    """템플릿별 이미지 추출용 짧은 지시."""
    ai = template.get("ai")
    if not isinstance(ai, dict):
        return None
    s = ai.get("extract")
    if s is None:
        return None
    t = str(s).strip()
    return t or None


def load_templates_manifest(base_dir: Path) -> List[Dict[str, Any]]:
    """
    병합 순서: xlsx/templates.json → hwpx/templates.json → (레거시) templates.json
    동일 id는 먼저 등록된 항목이 유지되고 이후는 무시(로그).
    xlsx·hwpx 매니페스트에서 kind 가 비어 있으면 각각 xlsx / hwpx 로 채움.
    """
    root = base_dir.resolve()
    merged: List[Dict[str, Any]] = []
    seen: set[str] = set()

    def _ingest(items: List[Dict[str, Any]], default_kind: Optional[str], source: str) -> None:
        for raw in items:
            tid = str(raw.get("id", "")).strip()
            if not tid:
                logger.warning("템플릿 id 가 비어 있어 건너뜀 (%s)", source)
                continue
            if tid in seen:
                logger.warning("템플릿 id 중복 — 이미 등록됨, 무시 (%s): %s", source, tid)
                continue
            t = dict(raw)
            if default_kind and not str(t.get("kind", "")).strip():
                t["kind"] = default_kind
            merged.append(t)
            seen.add(tid)

    _ingest(_read_templates_array_from_file(root / "xlsx" / "templates.json", "xlsx/templates.json"), "xlsx", "xlsx/templates.json")
    _ingest(_read_templates_array_from_file(root / "hwpx" / "templates.json", "hwpx/templates.json"), "hwpx", "hwpx/templates.json")
    legacy = root / "templates.json"
    if legacy.is_file():
        _ingest(_read_templates_array_from_file(legacy, "templates.json"), None, "templates.json(레거시)")

    if not merged:
        logger.warning("등록된 문서 템플릿이 없습니다. %s/xlsx|hwpx/templates.json 을 확인하세요.", base_dir)
    return merged


def get_template_by_id(base_dir: Path, template_id: str) -> Optional[Dict[str, Any]]:
    tid = (template_id or "").strip()
    for t in load_templates_manifest(base_dir):
        if str(t.get("id", "")).strip() == tid:
            return t
    return None


def resolve_template_file(base_dir: Path, template: Dict[str, Any]) -> Path:
    rel = (template.get("file") or "").strip().replace("\\", "/")
    if not rel or ".." in rel:
        raise ValueError("잘못된 템플릿 파일 경로입니다.")
    path = (base_dir / rel).resolve()
    root = base_dir.resolve()
    if root not in path.parents and path != root:
        raise ValueError("템플릿 파일이 허용된 디렉터리 밖입니다.")
    if not path.is_file():
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없습니다: {path.name}")
    return path


def fill_xlsx(template_path: Path, template: Dict[str, Any], values: Dict[str, Any]) -> bytes:
    """
    원본 경로를 openpyxl에 직접 넘기면 OneDrive/엑셀 점유 시 PermissionError 가 자주 난다.
    파일을 한 번에 읽어 BytesIO 로 열면 원본 핸들을 짧게만 쓴다.
    """
    try:
        raw = template_path.read_bytes()
    except PermissionError as e:
        raise PermissionError(
            f"엑셀 템플릿을 읽을 수 없습니다. "
            f"OneDrive 동기화 중이거나 엑셀에서 '{template_path.name}' 을 연 상태일 수 있습니다. "
            f"닫은 뒤 다시 시도하세요."
        ) from e
    except OSError as e:
        raise OSError(f"엑셀 템플릿 파일 읽기 실패: {template_path.name} ({e})") from e

    try:
        wb = load_workbook(io.BytesIO(raw))
    except Exception as e:
        logger.exception("openpyxl load_workbook 실패: %s", template_path)
        raise ValueError(
            f"엑셀 템플릿을 열 수 없습니다(손상·암호·지원 형식 확인): {template_path.name}"
        ) from e

    sheet_name = template.get("sheet")
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
    except KeyError as e:
        raise ValueError(f"시트 이름을 찾을 수 없습니다: {sheet_name!r}") from e

    _write_field_values_to_sheet(ws, template, values)

    bio = io.BytesIO()
    try:
        wb.save(bio)
    except Exception as e:
        logger.exception("openpyxl save 실패: %s", template_path)
        raise ValueError("엑셀 결과 저장에 실패했습니다. 템플릿에 보호·외부 링크 제한이 있는지 확인하세요.") from e
    return bio.getvalue()


def fill_hwpx(template_path: Path, template: Dict[str, Any], values: Dict[str, Any]) -> bytes:
    """HWPX(zip) 내부 모든 .xml 에서 {{field_id}} 또는 placeholder 를 치환."""
    out_bio = io.BytesIO()
    with zipfile.ZipFile(template_path, "r") as zin:
        with zipfile.ZipFile(out_bio, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename.lower().endswith(".xml"):
                    text = data.decode("utf-8", errors="surrogateescape")
                    for f in template.get("fields") or []:
                        if not isinstance(f, dict):
                            continue
                        fid = f.get("id")
                        if fid is None:
                            continue
                        token = (f.get("placeholder") or f"{{{{{fid}}}}}").strip()
                        repl = escape_xml_text(str(values.get(str(fid), "") or ""))
                        text = text.replace(token, repl)
                    data = text.encode("utf-8", errors="surrogateescape")
                zout.writestr(info, data)
    return out_bio.getvalue()


def normalize_table_columns(template: Dict[str, Any]) -> List[Dict[str, str]]:
    """extract_mode=table 용 table_columns 정규화."""
    raw = template.get("table_columns")
    if not isinstance(raw, list):
        return []
    out: List[Dict[str, str]] = []
    for c in raw:
        if not isinstance(c, dict) or c.get("id") is None:
            continue
        cid = str(c.get("id", "")).strip()
        if not cid:
            continue
        header = str(c.get("header", cid)).strip() or cid
        row: Dict[str, str] = {"id": cid, "header": header}
        col = c.get("column")
        if col is not None and str(col).strip():
            row["column"] = str(col).strip().upper()
        row["write"] = "false" if c.get("write") is False else "true"
        out.append(row)
    return out


def _safe_int(value: Any, default: int) -> int:
    try:
        n = int(value)
        return n if n > 0 else default
    except Exception:
        return default


def _normalize_row_key(v: Any) -> str:
    return "".join(str(v or "").split()).lower()


def _load_table_row_map(template: Dict[str, Any]) -> Dict[str, int]:
    raw = template.get("table_row_map")
    if not isinstance(raw, dict):
        return {}
    out: Dict[str, int] = {}
    for k, v in raw.items():
        nk = _normalize_row_key(k)
        if not nk:
            continue
        rv = _safe_int(v, -1)
        if rv > 0:
            out[nk] = rv
    return out


def _copy_row_styles(ws, src_row: int, dst_row: int, col_letters: List[str]) -> None:
    for col in col_letters:
        src = ws[f"{col}{src_row}"]
        dst = ws[f"{col}{dst_row}"]
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.protection:
            dst.protection = copy(src.protection)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def _write_field_values_to_sheet(ws, template: Dict[str, Any], values: Dict[str, Any]) -> None:
    for f in template.get("fields") or []:
        if not isinstance(f, dict):
            continue
        cell = f.get("cell")
        fid = f.get("id")
        if not cell or fid is None:
            continue
        try:
            ws[cell] = values.get(str(fid), "") or ""
        except Exception as e:
            logger.warning("셀 쓰기 실패 %s: %s", cell, e)
            raise ValueError(f"셀에 값을 넣을 수 없습니다: {cell} (병합·보호 시트 등 확인)") from e


def build_xlsx_table_bytes(
    template_path: Path,
    template: Dict[str, Any],
    rows: List[Dict[str, Any]],
    values: Optional[Dict[str, Any]] = None,
) -> bytes:
    """템플릿 xlsx를 열어 서식을 유지한 채 table_columns 기준으로 데이터 행을 채운다."""
    cols = normalize_table_columns(template)
    if not cols:
        raise ValueError("table_columns 가 비어 있습니다.")

    try:
        raw = template_path.read_bytes()
    except PermissionError as e:
        raise PermissionError(
            f"엑셀 템플릿을 읽을 수 없습니다. "
            f"OneDrive 동기화 중이거나 엑셀에서 '{template_path.name}' 을 연 상태일 수 있습니다. "
            f"닫은 뒤 다시 시도하세요."
        ) from e
    except OSError as e:
        raise OSError(f"엑셀 템플릿 파일 읽기 실패: {template_path.name} ({e})") from e

    try:
        wb = load_workbook(io.BytesIO(raw))
    except Exception as e:
        logger.exception("openpyxl load_workbook 실패(표 추출): %s", template_path)
        raise ValueError(
            f"엑셀 템플릿을 열 수 없습니다(손상·암호·지원 형식 확인): {template_path.name}"
        ) from e

    sheet_name = (template.get("sheet") or template.get("sheet_title") or "").strip()
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
    except KeyError as e:
        raise ValueError(f"시트 이름을 찾을 수 없습니다: {sheet_name!r}") from e

    if values:
        _write_field_values_to_sheet(ws, template, values)

    start_row = _safe_int(template.get("table_start_row"), 2)
    style_row = _safe_int(template.get("table_style_row"), start_row)
    clear_existing = bool(template.get("table_clear_existing", True))
    force_wrap = bool(template.get("table_force_wrap", True))
    row_key = str(template.get("table_row_key") or "").strip()
    row_map = _load_table_row_map(template)
    col_letters: List[str] = []
    col_writable: List[bool] = []
    for ci, c in enumerate(cols, start=1):
        col_ref = str(c.get("column") or get_column_letter(ci)).strip().upper()
        col_letters.append(col_ref)
        col_writable.append(str(c.get("write", "true")).lower() != "false")

    if clear_existing:
        min_clear = min(row_map.values()) if row_map else start_row
        max_clear = max(row_map.values()) if row_map else ws.max_row
        end_clear = max(max_clear, ws.max_row if not row_map else max_clear)
        for ri in range(min_clear, end_clear + 1):
            for col in col_letters:
                ws[f"{col}{ri}"].value = None

    ri = start_row
    for row in rows:
        if not isinstance(row, dict):
            continue
        if not any(str(row.get(c["id"], "") or "").strip() for c in cols):
            continue
        target_row = ri
        if row_key and row_map:
            mapped = row_map.get(_normalize_row_key(row.get(row_key)))
            if mapped:
                target_row = mapped
            else:
                # 팀 라벨 매핑이 필요한 템플릿에서 라벨이 없으면 오배치를 막기 위해 건너뜀.
                continue
        if target_row > ws.max_row and style_row > 0:
            _copy_row_styles(ws, style_row, target_row, col_letters)
        for ci, c in enumerate(cols, start=1):
            if not col_writable[ci - 1]:
                continue
            v = row.get(c["id"])
            col_ref = col_letters[ci - 1]
            cell = ws[f"{col_ref}{target_row}"]
            cell.value = "" if v is None else str(v)
            if force_wrap:
                base = cell.alignment or Alignment()
                new_align = copy(base)
                new_align.wrap_text = True
                cell.alignment = new_align
        if not row_map:
            ri += 1

    bio = io.BytesIO()
    try:
        wb.save(bio)
    except Exception as e:
        logger.exception("openpyxl save 실패(표 추출): %s", template_path)
        raise ValueError("엑셀 결과 저장에 실패했습니다. 템플릿에 보호·외부 링크 제한이 있는지 확인하세요.") from e
    return bio.getvalue()


def render_filled_document(base_dir: Path, template_id: str, values: Dict[str, Any]) -> tuple[bytes, str, str]:
    template = get_template_by_id(base_dir, template_id)
    if not template:
        raise ValueError("알 수 없는 템플릿입니다.")
    path = resolve_template_file(base_dir, template)
    kind = str(template.get("kind", "")).lower().strip()
    if kind == "xlsx":
        body = fill_xlsx(path, template, values)
        return body, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"
    if kind in ("hwpx", "hwp"):
        body = fill_hwpx(path, template, values)
        return body, "application/octet-stream", ".hwpx"
    raise ValueError(f"지원하지 않는 템플릿 종류입니다: {kind}")
