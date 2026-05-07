from pathlib import Path
from datetime import datetime
import hashlib
import io
from typing import Optional
from fastapi import APIRouter, UploadFile, File, HTTPException, Form, Depends
from fastapi.responses import Response
from openpyxl import Workbook
from app.services.vision_ai_service import VisionService
from app.core.config import settings
from app.db.repos.export import ExportRepository
from app.db.deps import get_export_repo
from app.core.auth import require_session

router = APIRouter(prefix="/api/vision", tags=["Vision"])
vision_svc = VisionService(api_key=settings.GEMINI_API_KEY)


def _base_storage() -> Path:
    return Path(settings.UPLOADS_DIR)


def _to_worklog_records(extracted: dict) -> list[dict]:
    work_code = str(extracted.get("project_code") or "").strip()
    input_date = str(extracted.get("work_date") or "").strip()
    regular_count = float(extracted.get("regular_count") or 0)
    daily_count = float(extracted.get("daily_count") or 0)
    signalman_count = float(extracted.get("signalman_count") or 0)
    excavator_6w = float(extracted.get("excavator_6w") or 0)
    excavator_3w = float(extracted.get("excavator_3w") or 0)
    dump_15t = float(extracted.get("dump_15t") or 0)
    crane_count = float(extracted.get("crane_count") or 0)
    connection_count = float(extracted.get("connection_count") or 0)

    is_night = bool(extracted.get("is_night"))

    def make_row(item_name: str, value: float) -> dict:
        value = float(value or 0)
        return {
            "work_code": work_code,
            "input_date": input_date,
            "item_name": item_name,
            "day_value": 0 if is_night else value,
            "night_value": value if is_night else 0,
        }

    return [
        make_row("상용직", regular_count),
        make_row("일용직", daily_count),
        make_row("모범신호수", signalman_count),
        make_row("6W", excavator_6w),
        make_row("3W", excavator_3w),
        make_row("덤프15T", dump_15t),
        make_row("크레인", crane_count),
        make_row("접속", connection_count),
    ]


def _to_worklog_structured_record(extracted: dict) -> dict:
    is_night = bool(extracted.get("is_night"))

    def split_day_night(value: float) -> tuple[float, float]:
        value = float(value or 0)
        return (0.0, value) if is_night else (value, 0.0)

    regular_day, regular_night = split_day_night(extracted.get("regular_count") or 0)
    daily_day, daily_night = split_day_night(extracted.get("daily_count") or 0)
    signalman_day, signalman_night = split_day_night(extracted.get("signalman_count") or 0)
    w6_day, w6_night = split_day_night(extracted.get("excavator_6w") or 0)
    w3_day, w3_night = split_day_night(extracted.get("excavator_3w") or 0)
    dump15t_day, dump15t_night = split_day_night(extracted.get("dump_15t") or 0)
    crane_day, crane_night = split_day_night(extracted.get("crane_count") or 0)
    connection_day, connection_night = split_day_night(extracted.get("connection_count") or 0)

    return {
        "work_code": str(extracted.get("project_code") or "").strip(),
        "input_date": str(extracted.get("work_date") or "").strip(),
        "regular_day": regular_day, "regular_night": regular_night,
        "daily_day": daily_day, "daily_night": daily_night,
        "signalman_day": signalman_day, "signalman_night": signalman_night,
        "w6_day": w6_day, "w6_night": w6_night,
        "w3_day": w3_day, "w3_night": w3_night,
        "dump15t_day": dump15t_day, "dump15t_night": dump15t_night,
        "crane_day": crane_day, "crane_night": crane_night,
        "watertruck_day": 0.0, "watertruck_night": 0.0,
        "mcm_day": 0.0, "mcm_night": 0.0,
        "connection_day": connection_day, "connection_night": connection_night,
        "outsource_1": 0.0, "outsource_2": 0.0,
    }


def _build_worklog_excel_bytes(structured_record: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    def num(v: object) -> float:
        try:
            return float(v or 0)
        except (TypeError, ValueError):
            return 0.0

    input_date = str(structured_record.get("input_date") or "").strip()
    parsed_date: object = input_date
    if input_date:
        try:
            parsed_date = datetime.strptime(input_date, "%Y-%m-%d").date()
        except ValueError:
            parsed_date = input_date

    rows = [
        ("지중No", str(structured_record.get("work_code") or "").strip()),
        ("투입일", parsed_date),
        ("상용직_주간", num(structured_record.get("regular_day"))),
        ("상용직_야간", num(structured_record.get("regular_night"))),
        ("일용직_주간", num(structured_record.get("daily_day"))),
        ("일용직_야간", num(structured_record.get("daily_night"))),
        ("모범 신호수_주간", num(structured_record.get("signalman_day"))),
        ("모범 신호수_야간", num(structured_record.get("signalman_night"))),
        ("6W_주간", num(structured_record.get("w6_day"))),
        ("6W_야간", num(structured_record.get("w6_night"))),
        ("3W_주간", num(structured_record.get("w3_day"))),
        ("3W_야간", num(structured_record.get("w3_night"))),
        ("덤프15T_주간", num(structured_record.get("dump15t_day"))),
        ("덤프15T_야간", num(structured_record.get("dump15t_night"))),
        ("크레인_주간", num(structured_record.get("crane_day"))),
        ("크레인_야간", num(structured_record.get("crane_night"))),
        ("물청소차_주간", num(structured_record.get("watertruck_day"))),
        ("물청소차_야간", num(structured_record.get("watertruck_night"))),
        ("MCM_주간", num(structured_record.get("mcm_day"))),
        ("MCM_야간", num(structured_record.get("mcm_night"))),
        ("접속_주간", num(structured_record.get("connection_day"))),
        ("접속_야간", num(structured_record.get("connection_night"))),
        ("외주1", num(structured_record.get("outsource_1"))),
        ("외주2", num(structured_record.get("outsource_2"))),
    ]

    for idx, (label, value) in enumerate(rows, start=1):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)

    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


@router.post("/extract-worklog")
async def extract_worklog(
    file: UploadFile = File(...),
    _user_session=Depends(require_session),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="이미지 파일이 비어 있습니다.")

    data = await vision_svc.analyze_document(content, file.content_type or "image/jpeg")
    if not data:
        raise HTTPException(status_code=500, detail="AI 분석 실패")

    records = _to_worklog_records(data)
    structured_record = _to_worklog_structured_record(data)
    return {
        "message": "추출 완료",
        "source": "ai",
        "raw": data,
        "structured_record": structured_record,
        "records": records,
    }


@router.post("/extract-worklog-xlsx")
async def extract_worklog_xlsx(
    file: UploadFile = File(...),
    _user_session=Depends(require_session),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="이미지 파일이 비어 있습니다.")

    data = await vision_svc.analyze_document(content, file.content_type or "image/jpeg")
    if not data:
        raise HTTPException(status_code=500, detail="AI 분석 실패")

    structured_record = _to_worklog_structured_record(data)
    blob = _build_worklog_excel_bytes(structured_record)
    return Response(
        content=blob,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="worklog_extract.xlsx"'},
    )


@router.post("/upload")
async def process_document(
    file: UploadFile = File(...),
    upload_category: str = Form(default=""),
    linked_schedule_id: Optional[int] = Form(default=None),
    note: str = Form(default=""),
    user_session=Depends(require_session),
    export_repo: ExportRepository = Depends(get_export_repo),
):
    content = await file.read()
    file_hash = hashlib.sha256(content).hexdigest()
    file_size = len(content)
    base_storage = _base_storage()

    if upload_category:
        date_dir = datetime.now().strftime("%Y-%m-%d")
        folder = base_storage / date_dir / upload_category
        folder.mkdir(parents=True, exist_ok=True)
        safe_name = datetime.now().strftime("%H%M%S")
        save_path = folder / f"{safe_name}{Path(file.filename).suffix}"
        with open(save_path, "wb") as f:
            f.write(content)
        upload_id = export_repo.save_photo_upload(
            category=upload_category,
            file_path=str(save_path),
            uploaded_by=user_session["user_id"],
            uploaded_device=user_session.get("device_name", "unknown-device"),
            related_date=date_dir,
            file_size=file_size,
            file_sha256=file_hash,
            linked_schedule_id=linked_schedule_id,
            note=note,
        )
        return {"message": "사진 저장 완료", "id": upload_id, "folder": str(folder), "filename": save_path.name}

    data = await vision_svc.analyze_document(content, file.content_type)

    if not data:
        raise HTTPException(status_code=500, detail="AI 분석 실패")

    date_dir = datetime.now().strftime("%Y-%m-%d")
    folder = base_storage / date_dir / data["doc_type"]
    folder.mkdir(parents=True, exist_ok=True)

    safe_name = f"{data['work_date']}_{data['project_name']}_{data['project_code']}".replace("/", "-")

    if data["doc_type"] == "작업일지":
        structured_record = _to_worklog_structured_record(data)
        excel_blob = _build_worklog_excel_bytes(structured_record)
        with open(folder / f"{safe_name}.xlsx", "wb") as xlsx_file:
            xlsx_file.write(excel_blob)

    with open(folder / f"{safe_name}{Path(file.filename).suffix}", "wb") as f:
        f.write(content)

    export_repo.save_photo_upload(
        category=data["doc_type"],
        file_path=str(folder / f"{safe_name}{Path(file.filename).suffix}"),
        uploaded_by=user_session["user_id"],
        uploaded_device=user_session.get("device_name", "unknown-device"),
        related_date=date_dir,
        file_size=file_size,
        file_sha256=file_hash,
    )

    return {"message": "분류 및 저장 완료", "folder": str(folder), "filename": safe_name}
